import pdfplumber
import re
import openpyxl
import streamlit as st

def extract_pdf_data(filepath):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Set column headers
    sheet['A1'] = 'AWB Number'
    sheet['B1'] = 'Date&Time'

    # Define regex patterns
    awb_regex = r"\b\d{14,15}\b"
    datetime_regex = r"\b\d{1,2}:\d{2}[AP]M [A-Z][a-z]{2} \d{2},\d{4}\b"

    # Open the PDF file
    with pdfplumber.open(filepath) as pdf:
        # Loop through all the pages in the PDF
        for page in pdf.pages:
            # Extract text from the page
            text = page.extract_text()
            
            # Extract AWB numbers from the page and write to the Excel sheet
            awb_numbers = re.findall(awb_regex, text)
            for i, awb in enumerate(awb_numbers):
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=awb)

            # Extract datetime from the page and write to the Excel sheet
            datetime = re.search(datetime_regex, text)
            if datetime:
                finalDate = datetime
        # Write datetime to all rows in Excel sheet
        for row in range(2, sheet.max_row+1):
            sheet.cell(row=row, column=2, value=finalDate.group(0))

    # Save the Excel workbook
    output_filepath = "output.xlsx"
    workbook.save(output_filepath)

    return output_filepath

st.title("PDF Text Extraction (Delhivery)")
st.write("This app extracts AWB numbers and date/time information from a PDF and writes the data to an Excel file.")
# Define custom CSS style

# Allow user to upload a file
file = st.file_uploader("Upload a PDF file", type="pdf")

# When the user clicks the "Extract" button, extract the data from the PDF and write to an Excel file
if st.button("Extract"):
    if file is not None:
        output_filepath = extract_pdf_data(file)
        st.write(f"Data extracted from {file.name} and written to {output_filepath}.")
    else:
        st.write("Please upload a PDF file.")

# Add download button for the output file
if 'output_filepath' in locals():
    with open(output_filepath, 'rb') as f:
        data = f.read()
    st.download_button(
        label="Download output file",
        data=data,
        file_name="output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
